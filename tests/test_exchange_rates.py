import json
from datetime import datetime

import pytest
import requests
from openpyxl import Workbook

import exchange_rates
from exchange_rates import format_rate, load_exchange_rates, parse_gdt_latest_rate


def build_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Parallel", None, None, None, None, "Official"])
    sheet.append([2025])
    sheet.append(["March", 4002.3, 4011.9, 4007.1, None, None, 4005])
    sheet.append(["January", 4018, 4032, 4025, None, None, 4024])
    sheet.append(["Annual 2025", None, None, None, None, None, 4050])
    sheet.append(["Annual 2010", None, None, None, None, None, 4200])
    workbook.properties.modified = datetime(2026, 9, 3, 16, 45)
    workbook.save(path)


def test_parser_supports_monthly_and_annual_only_years(tmp_path):
    path = tmp_path / "rates.xlsx"
    build_workbook(path)

    result = load_exchange_rates(path)

    assert list(result["years"]) == [2025, 2010]
    assert result["years"][2025]["annual"] is None
    assert [row["month"] for row in result["years"][2025]["months"]] == [
        "Jan",
        "Mar",
    ]
    assert result["years"][2010] == {
        "annual": 4200,
        "months": [],
        "toi_rate_available": False,
        "annual_method": "historical_average",
        "annual_source_url": None,
        "annual_published_at": None,
        "annual_verification": None,
    }


def test_parser_does_not_invent_toi_rate_before_fiscal_year_end(tmp_path):
    path = tmp_path / "rates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([2024])
    sheet.append(["January", 4000, 4010, 4005, None, None, 3990])
    sheet.append(["February", 4020, 4030, 4025, None, None, 4010])
    workbook.save(path)

    result = load_exchange_rates(path)

    assert result["years"][2024]["annual"] is None
    assert result["years"][2024]["toi_rate_available"] is False
    assert result["years"][2024]["annual_method"] == "unavailable"


def test_parser_does_not_use_december_monthly_rate_as_annual_toi(tmp_path):
    path = tmp_path / "rates.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([2024])
    sheet.append(["January", 4080, 4091, 4085.5, None, None, 4083])
    sheet.append(["December", 4024, 4035, 4029.5, None, None, 4025])
    workbook.save(path)

    result = load_exchange_rates(path)

    assert result["years"][2024]["annual"] is None
    assert result["years"][2024]["toi_rate_available"] is False
    assert result["years"][2024]["annual_method"] == "unavailable"


def test_parser_merges_verified_monthly_updates_without_duplicates(tmp_path):
    path = tmp_path / "rates.xlsx"
    build_workbook(path)
    updates = {
        "monthly_rates": [
            {
                "year": 2025,
                "month": "March",
                "purchase": 3996,
                "sale": 4008,
                "midpoint": 4002,
                "official": 4000,
            },
            {
                "year": 2026,
                "month": "January",
                "purchase": 4023,
                "sale": 4034,
                "midpoint": 4029,
                "official": 4026,
            },
        ]
    }
    path.with_name("exchange_rate_updates.json").write_text(
        json.dumps(updates), encoding="utf-8"
    )

    result = load_exchange_rates(path)

    months_2025 = result["years"][2025]["months"]
    assert [month["month"] for month in months_2025] == ["Jan", "Mar"]
    assert months_2025[1]["purchase"] == 3996
    assert result["years"][2026]["months"][0]["official"] == 4026
    assert result["years"][2026]["annual"] is None
    assert result["years"][2026]["toi_rate_available"] is False


def test_parser_does_not_use_historical_average_for_2021_toi(tmp_path):
    path = tmp_path / "rates.xlsx"
    build_workbook(path)
    path.with_name("exchange_rate_updates.json").write_text(
        json.dumps(
            {
                "historical_annual_average_rates": [
                    {"year": 2021, "rate": 4098.72279505888}
                ]
            }
        ),
        encoding="utf-8",
    )

    result = load_exchange_rates(path)

    assert result["years"][2021]["annual"] is None
    assert result["years"][2021]["annual_method"] == "unavailable"


def test_rate_formatting_uses_whole_numbers_and_thousands_separators():
    assert format_rate(4002.3) == "4,002"
    assert format_rate(4011.9) == "4,012"
    assert format_rate(4050.0) == "4,050"


def test_gdt_parser_reads_first_official_usd_rate():
    page = """
        <table><tr><th>Release Date</th><th>Symbol</th><th>Official Rate</th></tr>
        <tr><td>September 2, 2026</td><td>USD/KHR</td><td>4,047</td>
        <td>National Bank of Cambodia</td></tr>
        <tr><td>September 1, 2026</td><td>USD/KHR</td><td>4,046</td>
        <td>National Bank of Cambodia</td></tr></table>
    """

    latest = parse_gdt_latest_rate(page)

    assert latest["rate"] == 4047
    assert latest["published_at"] == datetime(2026, 9, 2)


def test_sourced_closing_rate_overrides_unverified_workbook_annual(tmp_path):
    path = tmp_path / "rates.xlsx"
    build_workbook(path)
    source = {"year": 2025, "rate": 4013, "published_at": "2025-12-31",
              "source_url": "https://www.tax.gov.kh/en/exchange-rate?for_year=2025&for_month=12"}
    path.with_name("exchange_rate_updates.json").write_text(
        json.dumps({"annual_closing_rates": [source]}), encoding="utf-8"
    )
    record = load_exchange_rates(path)["years"][2025]
    assert record["annual"] == 4013
    assert record["toi_rate_available"] is True
    assert record["annual_source_url"] == source["source_url"]
    assert record["annual_published_at"] == "2025-12-31"


@pytest.mark.parametrize("change", [
    {"published_at": "2024-12-31"}, {"published_at": "2025-11-30"},
    {"source_url": "https://example.com/rates"}, {"rate": -1}, {"rate": "NaN"},
])
def test_rejects_invalid_annual_source(tmp_path, change):
    path = tmp_path / "rates.xlsx"
    build_workbook(path)
    source = {"year": 2025, "rate": 4013, "published_at": "2025-12-31",
              "source_url": "https://www.tax.gov.kh/en/exchange-rate", **change}
    path.with_name("exchange_rate_updates.json").write_text(
        json.dumps({"annual_closing_rates": [source]}), encoding="utf-8"
    )
    with pytest.raises(ValueError):
        load_exchange_rates(path)


@pytest.fixture
def gdt_fetch(monkeypatch):
    monkeypatch.setattr(exchange_rates, "_gdt_cache", {
        "expires_at": 0.0, "value": None, "checked_at": None,
        "attempted_at": None, "stale": False,
    })
    state = {"now": 100.0, "calls": 0, "fail": False, "malformed": False}
    monkeypatch.setattr(exchange_rates.time, "monotonic", lambda: state["now"])

    def get(*args, **kwargs):
        state["calls"] += 1
        if state["fail"]:
            raise requests.Timeout("test outage")
        response = requests.Response()
        response.status_code = 200
        response._content = (
            b"unrecognized response" if state["malformed"] else
            b"<td>September 4, 2026</td><td>USD/KHR</td><td>4048</td><td>National Bank of Cambodia</td>"
        )
        return response

    monkeypatch.setattr(exchange_rates.requests, "get", get)
    return state


def test_refresh_bypasses_cache_but_coalesces_repeated_clicks(gdt_fetch):
    first = exchange_rates.fetch_latest_gdt_rate()
    assert first["value"]["rate"] == 4048
    assert first["cached"] is False
    assert exchange_rates.fetch_latest_gdt_rate(force=True)["refresh_throttled"] is True
    assert gdt_fetch["calls"] == 1
    gdt_fetch["now"] += 31
    assert exchange_rates.fetch_latest_gdt_rate()["cached"] is True
    refreshed = exchange_rates.fetch_latest_gdt_rate(force=True)
    assert refreshed["cached"] is False
    assert gdt_fetch["calls"] == 2


@pytest.mark.parametrize("failure", ["fail", "malformed"])
def test_failed_check_preserves_rate_and_marks_it_stale_then_recovers(gdt_fetch, failure):
    first = exchange_rates.fetch_latest_gdt_rate()
    gdt_fetch["now"] += 31
    gdt_fetch[failure] = True
    failed = exchange_rates.fetch_latest_gdt_rate(force=True)
    assert failed["value"] == first["value"]
    assert failed["checked_at"] == first["checked_at"]
    assert failed["stale"] is True
    assert failed["attempted_at"] >= first["attempted_at"]
    assert exchange_rates.fetch_latest_gdt_rate()["stale"] is True
    gdt_fetch["now"] += 31
    gdt_fetch[failure] = False
    assert exchange_rates.fetch_latest_gdt_rate(force=True)["stale"] is False


def test_initial_outage_has_no_rate_and_retries_after_failure_ttl(gdt_fetch):
    gdt_fetch["fail"] = True
    result = exchange_rates.fetch_latest_gdt_rate()
    assert result["value"] is None
    assert result["checked_at"] is None
    assert result["stale"] is True
    gdt_fetch["now"] += 601
    gdt_fetch["fail"] = False
    assert exchange_rates.fetch_latest_gdt_rate()["value"]["rate"] == 4048


def test_bundled_rates_have_separate_year_end_sources():
    from pathlib import Path
    result = load_exchange_rates(Path(__file__).resolve().parents[1] / "data" / "Exchange Rate.xlsx")
    for year, rate in {2022: 4117, 2023: 4085, 2024: 4025, 2025: 4013}.items():
        record = result["years"][year]
        assert record["annual"] == rate
        assert record["annual_published_at"].startswith(f"{year}-12-")
        assert f"for_year={year}" in record["annual_source_url"]
    assert result["years"][2026]["annual"] is None
    assert result["years"][2013]["toi_rate_available"] is False


@pytest.mark.parametrize("year, rate, method", [
    (2014, 4038, "gdt_annual_average"), (2015, 4060, "gdt_annual_average"),
    (2016, 4037, "gdt_annual_average"), (2017, 4045, "gdt_annual_average"),
    (2018, 4045, "gdt_annual_average"), (2019, 4052, "gdt_annual_average"),
    (2020, 4045, "gdt_year_end"), (2021, 4074, "gdt_year_end"),
])
def test_bundled_reported_toi_rates_replace_historical_averages(year, rate, method):
    from pathlib import Path
    result = load_exchange_rates(Path(__file__).resolve().parents[1] / "data" / "Exchange Rate.xlsx")
    record = result["years"][year]
    assert record["annual"] == rate
    assert record["annual_method"] == method
    assert record["annual_verification"] == "financial_report"
    assert record["annual_published_at"] is None
    assert record["annual_source_url"].endswith(".pdf")
    assert record["toi_rate_available"] is True


def test_direct_official_source_takes_precedence_over_report(tmp_path):
    path = tmp_path / "rates.xlsx"
    build_workbook(path)
    path.with_name("exchange_rate_updates.json").write_text(json.dumps({
        "annual_closing_rates": [{"year": 2020, "rate": 4045,
            "published_at": "2020-12-31", "source_url": "https://www.tax.gov.kh/en/exchange-rate"}],
        "reported_annual_toi_rates": [{"year": 2020, "rate": 4000,
            "annual_method": "gdt_year_end", "verification": "financial_report",
            "source_url": "https://example.com/report.pdf"}],
    }), encoding="utf-8")
    record = load_exchange_rates(path)["years"][2020]
    assert record["annual"] == 4045
    assert record["annual_verification"] == "official_publication"
