"""Read exchange-rate records from the supplied NBC workbook."""

from __future__ import annotations

import html
import json
import re
import threading
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
import requests


MONTHS = {
    "january": (1, "Jan"),
    "february": (2, "Feb"),
    "march": (3, "Mar"),
    "april": (4, "Apr"),
    "may": (5, "May"),
    "june": (6, "Jun"),
    "july": (7, "Jul"),
    "august": (8, "Aug"),
    "september": (9, "Sep"),
    "october": (10, "Oct"),
    "november": (11, "Nov"),
    "december": (12, "Dec"),
}

GDT_EXCHANGE_RATE_URL = "https://www.tax.gov.kh/gdtwebsiteweb/en/exchange-rate"
GDT_CACHE_SECONDS = 6 * 60 * 60
GDT_TOI_CLOSING_RATE_START_YEAR = 2022
_gdt_cache = {"expires_at": 0.0, "value": None}
_gdt_cache_lock = threading.Lock()


def _number(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", "").strip())
        except ValueError:
            return None
    return None


def _year(value):
    number = _number(value)
    if number is not None and number.is_integer() and 1900 <= number <= 2200:
        return int(number)
    return None


def _annual_year(value):
    if not isinstance(value, str) or "annual" not in value.casefold():
        return None
    for part in value.replace("/", " ").replace("-", " ").split():
        if part.isdigit() and 1900 <= int(part) <= 2200:
            return int(part)
    return None


def load_exchange_rates(path: str | Path) -> dict:
    """Return annual and available monthly rates without assuming fixed rows."""
    path = Path(path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    years: dict[int, dict] = {}

    for sheet in workbook.worksheets:
        current_year = None
        for values in sheet.iter_rows(values_only=True):
            first = values[0] if values else None
            marker_year = _year(first)
            if marker_year:
                current_year = marker_year
                years.setdefault(marker_year, {"annual": None, "months": []})
                continue

            annual_year = next(
                (_annual_year(value) for value in values if _annual_year(value)),
                None,
            )
            if annual_year:
                numbers = [_number(value) for value in values]
                annual_value = next(
                    (number for number in reversed(numbers) if number is not None),
                    None,
                )
                if annual_value is not None:
                    years.setdefault(annual_year, {"annual": None, "months": []})[
                        "annual"
                    ] = annual_value
                continue

            month_key = str(first).strip().casefold() if first is not None else ""
            if current_year is None or month_key not in MONTHS:
                continue

            purchase = _number(values[1]) if len(values) > 1 else None
            sale = _number(values[2]) if len(values) > 2 else None
            midpoint = _number(values[3]) if len(values) > 3 else None
            # NBC uses purchase/sale/midpoint in older blocks and one rate in G
            # from 2012 onward. G is the comparable official midpoint in both.
            official = _number(values[6]) if len(values) > 6 else None
            if purchase is None or sale is None or midpoint is None or official is None:
                continue
            month_number, month_label = MONTHS[month_key]
            years[current_year]["months"].append(
                {
                    "number": month_number,
                    "month": month_label,
                    "purchase": purchase,
                    "sale": sale,
                    "midpoint": midpoint,
                    "official": official,
                }
            )

    modified = workbook.properties.modified
    workbook.close()

    updates_path = path.with_name("exchange_rate_updates.json")
    if updates_path.exists():
        updates = json.loads(updates_path.read_text(encoding="utf-8"))
        for annual_rate in updates.get("historical_annual_average_rates", []):
            year = int(annual_rate["year"])
            if year < GDT_TOI_CLOSING_RATE_START_YEAR:
                years.setdefault(year, {"annual": None, "months": []})["annual"] = float(
                    annual_rate["rate"]
                )
        for update in updates.get("monthly_rates", []):
            year = int(update["year"])
            record = years.setdefault(year, {"annual": None, "months": []})
            month_key = str(update["month"]).strip().casefold()
            if month_key not in MONTHS:
                continue
            month_number, month_label = MONTHS[month_key]
            month_record = {
                "number": month_number,
                "month": month_label,
                "purchase": float(update["purchase"]),
                "sale": float(update["sale"]),
                "midpoint": float(update["midpoint"]),
                "official": float(update["official"]),
            }
            record["months"] = [
                month for month in record["months"] if month["number"] != month_number
            ]
            record["months"].append(month_record)

    for year, record in years.items():
        record["months"].sort(key=lambda month: month["number"])
        december = next(
            (month for month in record["months"] if month["number"] == 12), None
        )
        if (
            record["annual"] is None
            and year >= GDT_TOI_CLOSING_RATE_START_YEAR
            and december
        ):
            record["annual"] = december["official"]
        record["toi_rate_available"] = record["annual"] is not None
        record["annual_method"] = (
            "historical_average"
            if record["annual"] is not None and year < GDT_TOI_CLOSING_RATE_START_YEAR
            else "gdt_year_end"
            if record["annual"] is not None
            else "unavailable"
        )

    return {
        "years": dict(sorted(years.items(), reverse=True)),
        "last_updated": modified if isinstance(modified, datetime) else None,
    }


def format_rate(value) -> str:
    """Format with Excel-style half-up rounding and thousands separators."""
    if value is None:
        return "—"
    return f"{int(float(value) + 0.5):,}"


def parse_gdt_latest_rate(page_html: str) -> dict | None:
    """Extract the first published USD/KHR row from the GDT page."""
    text = re.sub(r"<[^>]+>", " ", page_html)
    text = re.sub(r"\s+", " ", html.unescape(text))
    match = re.search(
        r"([A-Z][a-z]+\s+\d{1,2},\s+\d{4})\s+USD/KHR\s+([\d,]+)\s+"
        r"National Bank of Cambodia",
        text,
    )
    if not match:
        return None
    try:
        published_at = datetime.strptime(match.group(1), "%B %d, %Y")
        rate = int(match.group(2).replace(",", ""))
    except ValueError:
        return None
    return {"rate": rate, "published_at": published_at, "source_url": GDT_EXCHANGE_RATE_URL}


def fetch_latest_gdt_rate(force: bool = False) -> dict | None:
    """Fetch GDT's latest official rate, cached to avoid excessive requests."""
    now = time.monotonic()
    with _gdt_cache_lock:
        if not force and now < _gdt_cache["expires_at"]:
            return _gdt_cache["value"]

        try:
            response = requests.get(
                GDT_EXCHANGE_RATE_URL,
                headers={"User-Agent": "B03-Meeting-Room-Mini-App/1.0"},
                timeout=8,
            )
            response.raise_for_status()
            latest = parse_gdt_latest_rate(response.text)
        except requests.RequestException:
            latest = None

        if latest:
            _gdt_cache["value"] = latest
            _gdt_cache["expires_at"] = now + GDT_CACHE_SECONDS
        else:
            # Preserve the last known official rate during a temporary outage.
            _gdt_cache["expires_at"] = now + 10 * 60
        return _gdt_cache["value"]
